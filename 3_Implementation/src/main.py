# using pandas library
import pandas as pd
# using openpyxl to write in master sheet
from openpyxl import load_workbook


# Reading Excel File Using Pandas Library
# ########################SANGAM Function Starts###################################################################################################################
def sangam():
    san = pd.ExcelFile('PythonExcelSheets.xlsx')
    df1 = pd.read_excel(san, 'Sheet1')
    df2 = pd.read_excel(san, 'Sheet2')
    df3 = pd.read_excel(san, 'Sheet3')
    df4 = pd.read_excel(san, 'Sheet4')
    df5 = pd.read_excel(san, 'Sheet5')
    ps = pd.DataFrame(df1, columns=['Ps No', 'Name'])
    # print(ps)
    # converting ps no data frame to a List
    psnolist = ps['Ps No'].values.tolist()
    # Here Making a Dictionary From Priviously Created Ps No List
    dict1 = {"key1": psnolist}
    # printing Dictionary so that user can pick anyone Unique key from dictionary
    print(dict1)
    print("Enter Unique Ps No Of Person Unique (Keys Is In Dictionary Key1)")
    # Taking User Input For Unique Key
    ui = int(input())
    # This Part Will Check Wether User Has Inputed valid Key Or not
    key_to_lookup = 'key1'
    # if key is valid it will Print (Unique Key Is True) This Message
    if key_to_lookup in dict1:
        print("Unique Key Is True")
    # if key is valid it will Print (Please Enter Valid Unique Key of a Person) This Message
    else:
        print("Please Enter Valid Unique Key of a Person")
    # Now Here Locking The found Column which user has inputed so it will fetch data from row only infront of column
    df1.set_index("Ps No", inplace=True)
    result1 = df1.loc[ui]
    df2.set_index("Ps No", inplace=True)
    result2 = df2.loc[ui]
    df3.set_index("Ps No", inplace=True)
    result3 = df3.loc[ui]
    df4.set_index("Ps No", inplace=True)
    result4 = df4.loc[ui]
    df5.set_index("Ps No", inplace=True)
    result5 = df5.loc[ui]
    # creating a list to append data frames to a list
    ls = []
    lm = []
    # appending the list using for loop
    for i in range(len(result1)):
        ls.append(result1[i])
    lm.append(len(ls))
    for i in range(len(result2)):
        # appending the list but not repeating common items by this If Condition
        if i >= 2:
            ls.append(result2[i])
    lm.append(len(ls) + 2)
    for i in range(len(result3)):
        if i >= 2:
            ls.append(result3[i])
    lm.append(len(ls) + 2)
    for i in range(len(result4)):
        if i >= 2:
            ls.append(result4[i])
    lm.append(len(ls) + 2)
    for i in range(len(result5)):
        if i >= 2:
            ls.append(result5[i])
    lm.append(len(ls) + 2)
    # calculating length of complete list and assigning it to length variable
    # length = (len(ls))
    # print(length)
    print(ls)
    print(lm)
    summary = pd.DataFrame(lm)
    # opening Excel Workbook#######################################################################################################################################
    # path = "PythonExcelSheets.xlsx"
    ExcelWorkbook1 = load_workbook('PythonExcelSheets.xlsx')
    if 'summarysheet' not in ExcelWorkbook1.sheetnames:
        with pd.ExcelWriter('PythonExcelSheets.xlsx', engine="openpyxl", mode='a') as writer1:
            summary.to_excel(writer1, sheet_name='summarysheet')
    # return ls,lm
    return ls


# #################################SANGAM Function Ends############################################################################################################
# creating an Empty List
lk = []
print("How Much person Data you Want")
z = int(input())
for j in range(z):
    lk.append(sangam())
print(lk)
# converting my final list to a Data Frame
final = pd.DataFrame(lk)
path = "PythonExcelSheets.xlsx"
ExcelWorkbook = load_workbook('PythonExcelSheets.xlsx')
writer = pd.ExcelWriter('PythonExcelSheets.xlsx', engine='openpyxl')
writer.book = ExcelWorkbook
# checking if master sheet is already prsent then it will not create the master sheet again
if 'mastersheet' in ExcelWorkbook.sheetnames:
    pfd = ExcelWorkbook['mastersheet']
    ExcelWorkbook.remove(pfd)
final.to_excel(writer, sheet_name='mastersheet')
writer.save()
writer.close()
print(len(lk))
